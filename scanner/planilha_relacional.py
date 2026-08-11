"""
Estrutura profissional da planilha: modelo relacional com 4 abas.
1. FUNCIONÁRIOS (base oficial)  2. TREINAMENTOS  3. PRESENÇA (cresce no tempo)  4. DASHBOARD
"""

import os
import pandas as pd
from typing import Optional, List
from datetime import datetime
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO_PADRAO = os.path.join(_SCRIPT_DIR, "controle_treinamentos.xlsx")


def _carregar_excel_todas_abas(caminho: str) -> dict:
    """Carrega todas as abas do Excel. Retorna {nome_aba: DataFrame}."""
    if not os.path.exists(caminho):
        return {}
    try:
        return pd.read_excel(caminho, sheet_name=None)
    except Exception:
        return {}


def criar_estrutura_profissional(caminho_excel: str = ARQUIVO_PADRAO) -> str:
    """
    Cria ou mantém as 4 abas: FUNCIONÁRIOS, TREINAMENTOS, PRESENÇA, DASHBOARD.
    Se o arquivo já existir, preserva as outras abas e só garante que essas 4 existam.
    """
    colunas_funcionarios = ["Matrícula", "Nome", "Setor", "Ativo"]
    colunas_treinamentos = ["ID", "Nome"]  # ID sem espaço (ex: FO0603060006) para evitar erro
    colunas_presenca = ["Data", "ID_Treinamento", "Matrícula", "Status"]
    colunas_dashboard = ["Indicador", "Valor", "Detalhe"]

    existente = _carregar_excel_todas_abas(caminho_excel)
    sheets = {}

    # 1. FUNCIONÁRIOS
    if "FUNCIONÁRIOS" in existente and not existente["FUNCIONÁRIOS"].empty:
        sheets["FUNCIONÁRIOS"] = existente["FUNCIONÁRIOS"]
    else:
        df = pd.DataFrame(columns=colunas_funcionarios)
        df.loc[0] = [58974, "João", "Produção", "Sim"]
        df.loc[1] = [105822, "Maria", "Qualidade", "Sim"]
        sheets["FUNCIONÁRIOS"] = df

    # 2. TREINAMENTOS
    if "TREINAMENTOS" in existente and not existente["TREINAMENTOS"].empty:
        sheets["TREINAMENTOS"] = existente["TREINAMENTOS"]
    else:
        df = pd.DataFrame(columns=colunas_treinamentos)
        df.loc[0] = ["FO0603060006", "FO 060 306-0006"]
        df.loc[1] = ["FO0603060011", "FO 060 306-0011"]
        df.loc[2] = ["FO0603060016", "FO 060 306-0016"]
        sheets["TREINAMENTOS"] = df

    # 3. PRESENÇA
    if "PRESENÇA" in existente and not existente["PRESENÇA"].empty:
        sheets["PRESENÇA"] = existente["PRESENÇA"]
    else:
        df = pd.DataFrame(columns=colunas_presenca)
        df.loc[0] = ["25/02/2026", "FO0603060006", 140398, "OK"]
        sheets["PRESENÇA"] = df

    # 4. ListaScanner (onde o Excel importa o CSV gerado pelo scanner — não editar pelo exe)
    if "ListaScanner" in existente and not existente["ListaScanner"].empty:
        sheets["ListaScanner"] = existente["ListaScanner"]
    else:
        df = pd.DataFrame(columns=["Nome_Treinamento", "Matrícula"])
        df.loc[0] = ["(importe aqui o CSV do scanner)", ""]
        sheets["ListaScanner"] = df

    # 5. DASHBOARD
    if "DASHBOARD" in existente and not existente["DASHBOARD"].empty:
        sheets["DASHBOARD"] = existente["DASHBOARD"]
    else:
        df = pd.DataFrame(columns=colunas_dashboard)
        df.loc[0] = ["% Presença geral", "", "Fórmula ou tabela dinâmica"]
        df.loc[1] = ["Pendentes", "", "Quem falta treinamentos obrigatórios"]
        df.loc[2] = ["Concluíram todos", "", "Por setor"]
        sheets["DASHBOARD"] = df

    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        for nome, df in sheets.items():
            df.to_excel(writer, sheet_name=nome, index=False)
        wb = writer.book
        # MATRIZ_CONTROLE: Matrícula, Nome, depois uma coluna por ID (PRESENÇA = banco oficial)
        # Fórmula: =SE(CONT.SES(PRESENÇA!B:B;C$1;PRESENÇA!C:C;$A2)>0;"OK";"")
        df_treinos = sheets["TREINAMENTOS"]
        ids_treino = df_treinos["ID"].astype(str).tolist() if "ID" in df_treinos.columns and not df_treinos.empty else []
        df_func = sheets["FUNCIONÁRIOS"]
        ws_matriz = wb.create_sheet("MATRIZ_CONTROLE", len(wb.worksheets))
        ws_matriz["A1"] = "Matrícula"
        ws_matriz["B1"] = "Nome"
        for col_idx, id_t in enumerate(ids_treino, start=3):
            ws_matriz.cell(row=1, column=col_idx, value=id_t)
        for idx, row in df_func.iterrows():
            r = idx + 2
            ws_matriz.cell(row=r, column=1, value=row.get("Matrícula", ""))
            ws_matriz.cell(row=r, column=2, value=row.get("Nome", ""))
            for col_idx in range(3, 3 + len(ids_treino)):
                col_letter = get_column_letter(col_idx)
                ws_matriz.cell(
                    row=r, column=col_idx,
                    value=f'=SE(CONT.SES(PRESENÇA!B:B;{col_letter}$1;PRESENÇA!C:C;$A{r})>0;"OK";"")',
                )
        # Aba "Por Treinamento": nomes na linha 1, fórmulas puxam matrículas da ListaScanner
        ws = wb.create_sheet("Por Treinamento", len(wb.worksheets))
        ws["A1"] = "5s"
        ws["B1"] = "POLITICA GESTÃO"
        ws["C1"] = "ISO9001"
        # Arraste A2:C2 para baixo para listar mais matrículas por treinamento
        ws["A2"] = '=SEERRO(ÍNDICE(ListaScanner!$B:$B;MENOR(SE(ListaScanner!$A:$A=A$1;LIN(ListaScanner!$A:$A));LINHA()-1));"")'
        ws["B2"] = '=SEERRO(ÍNDICE(ListaScanner!$B:$B;MENOR(SE(ListaScanner!$A:$A=B$1;LIN(ListaScanner!$A:$A));LINHA()-1));"")'
        ws["C2"] = '=SEERRO(ÍNDICE(ListaScanner!$B:$B;MENOR(SE(ListaScanner!$A:$A=C$1;LIN(ListaScanner!$A:$A));LINHA()-1));"")'
        # Aba "Status por treinamento": uma linha por funcionário; colunas = treinamentos (Feito/Falta)
        ws_status = wb.create_sheet("Status por treinamento", len(wb.worksheets))
        ws_status["A1"] = "Matrícula"
        ws_status["B1"] = "Nome"
        ws_status["C1"] = "5s"
        ws_status["D1"] = "POLITICA GESTÃO"
        ws_status["E1"] = "ISO9001"
        ws_status["F1"] = "Treinamentos que faltam"
        df_func = sheets["FUNCIONÁRIOS"]
        for idx, row in df_func.iterrows():
            r = idx + 2
            ws_status.cell(row=r, column=1, value=row.get("Matrícula", ""))
            ws_status.cell(row=r, column=2, value=row.get("Nome", ""))
            for col, col_letter in [(3, "C"), (4, "D"), (5, "E")]:
                ws_status.cell(
                    row=r, column=col,
                    value=f'=SE(CONT.SES(ListaScanner!$A:$A;{col_letter}$1;ListaScanner!$B:$B;$A{r})=0;"Falta";"Feito")',
                )
            # Nome dos treinamentos que faltam (Excel 365: TEXTOJUNTAR; senão deixa em branco)
            ws_status.cell(
                row=r, column=6,
                value=f'=TEXTOJUNTAR(", ";VERDADEIRO;SE(C{r}="Falta";C$1;"");SE(D{r}="Falta";D$1;"");SE(E{r}="Falta";E$1;""))',
            )
        # Aba "Faltam por treinamento": por coluna (treinamento), lista de matrículas que faltam
        ws_faltam = wb.create_sheet("Faltam por treinamento", len(wb.worksheets))
        ws_faltam["A1"] = "5s"
        ws_faltam["B1"] = "POLITICA GESTÃO"
        ws_faltam["C1"] = "ISO9001"
        ws_faltam["A2"] = '=SEERRO(ÍNDICE(\'Status por treinamento\'!$A:$A;MENOR(SE(\'Status por treinamento\'!C:C="Falta";LIN(\'Status por treinamento\'!$A:$A);9^9);LINHA()-1));"")'
        ws_faltam["B2"] = '=SEERRO(ÍNDICE(\'Status por treinamento\'!$A:$A;MENOR(SE(\'Status por treinamento\'!D:D="Falta";LIN(\'Status por treinamento\'!$A:$A);9^9);LINHA()-1));"")'
        ws_faltam["C2"] = '=SEERRO(ÍNDICE(\'Status por treinamento\'!$A:$A;MENOR(SE(\'Status por treinamento\'!E:E="Falta";LIN(\'Status por treinamento\'!$A:$A);9^9);LINHA()-1));"")'
    return caminho_excel


def ler_funcionarios(caminho_excel: str = ARQUIVO_PADRAO) -> pd.DataFrame:
    """Lê a aba FUNCIONÁRIOS."""
    if not os.path.exists(caminho_excel):
        return pd.DataFrame()
    try:
        return pd.read_excel(caminho_excel, sheet_name="FUNCIONÁRIOS")
    except Exception:
        return pd.DataFrame()


def ler_treinamentos(caminho_excel: str = ARQUIVO_PADRAO) -> pd.DataFrame:
    """Lê a aba TREINAMENTOS."""
    if not os.path.exists(caminho_excel):
        return pd.DataFrame()
    try:
        return pd.read_excel(caminho_excel, sheet_name="TREINAMENTOS")
    except Exception:
        return pd.DataFrame()


def obter_nome_treinamento(caminho_excel: str, id_treinamento: str) -> Optional[str]:
    """
    Retorna o Nome da aba TREINAMENTOS para o ID dado (colunas ID e Nome, ou ID_Treinamento e Nome_Treinamento).
    Se não encontrar, retorna None.
    """
    df = ler_treinamentos(caminho_excel)
    if df.empty:
        return None
    id_col = "ID" if "ID" in df.columns else ("ID_Treinamento" if "ID_Treinamento" in df.columns else None)
    nome_col = "Nome" if "Nome" in df.columns else ("Nome_Treinamento" if "Nome_Treinamento" in df.columns else None)
    if not id_col or not nome_col:
        return None
    id_str = str(id_treinamento).strip()
    linhas = df[df[id_col].astype(str).str.strip() == id_str]
    if linhas.empty:
        return None
    nome = linhas.iloc[0][nome_col]
    return str(nome).strip() if pd.notna(nome) and str(nome).strip() else None


def ler_presenca(caminho_excel: str = ARQUIVO_PADRAO) -> pd.DataFrame:
    """Lê a aba PRESENÇA."""
    if not os.path.exists(caminho_excel):
        return pd.DataFrame()
    try:
        return pd.read_excel(caminho_excel, sheet_name="PRESENÇA")
    except Exception:
        return pd.DataFrame()


def adicionar_presenca_ocr(
    data: str,
    id_treinamento: str,
    matriculas_presentes: List[str],
    caminho_excel: str = ARQUIVO_PADRAO,
) -> str:
    """
    Compara com a base FUNCIONÁRIOS e adiciona linhas na PRESENÇA: Presente ou Ausente.
    matriculas_presentes = lista de matrículas lidas pelo OCR (quem assinou).
    Para cada funcionário na base: se matrícula em matriculas_presentes → Status=Presente, senão Ausente.
    """
    func = ler_funcionarios(caminho_excel)
    if func.empty or "Matrícula" not in func.columns:
        # Sem base: só registra os presentes
        pres = pd.DataFrame({
            "Data": [data] * len(matriculas_presentes),
            "ID_Treinamento": [id_treinamento] * len(matriculas_presentes),
            "Matrícula": [str(m).strip() for m in matriculas_presentes],
            "Status": ["Presente"] * len(matriculas_presentes),
        })
    else:
        mat_base = func["Matrícula"].astype(str).str.strip().tolist()
        presentes_set = {str(m).strip() for m in matriculas_presentes}
        linhas = []
        for mat in mat_base:
            linhas.append({
                "Data": data,
                "ID_Treinamento": id_treinamento,
                "Matrícula": mat,
                "Status": "Presente" if mat in presentes_set else "Ausente",
            })
        pres = pd.DataFrame(linhas)
    # Append à aba PRESENÇA
    existente = ler_presenca(caminho_excel)
    if existente.empty:
        pres_final = pres
    else:
        pres_final = pd.concat([existente, pres], ignore_index=True)
    # Salvar todas as abas (preservar as outras, só atualizar PRESENÇA)
    todas = _carregar_excel_todas_abas(caminho_excel)
    if not todas:
        todas = {}
    todas["PRESENÇA"] = pres_final
    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        for nome, df in todas.items():
            df.to_excel(writer, sheet_name=nome[:31], index=False)
    return caminho_excel


def gerar_csv_presenca(
    data: str,
    id_treinamento: str,
    matriculas_presentes: List[str],
    nome_treinamento: Optional[str] = None,
    nomes_presentes: Optional[List[str]] = None,
    pasta_saida: Optional[str] = None,
) -> str:
    """
    Gera CSV para o Excel importar. O scanner puxa apenas: matrículas (quem assinou) e nome do treinamento.
    2 colunas: Nome_Treinamento, Matrícula. O resto você faz manual no Excel.
    """
    data_arq = data.replace("/", "-").replace(".", "-")
    nome_arq = f"presenca_{id_treinamento}_{data_arq}.csv"
    pasta = pasta_saida or _SCRIPT_DIR
    caminho = os.path.join(pasta, nome_arq)
    rotulo = (nome_treinamento or id_treinamento).strip()
    nomes = nomes_presentes or [""] * len(matriculas_presentes)
    nomes = (nomes + [""] * len(matriculas_presentes))[:len(matriculas_presentes)]
    df = pd.DataFrame({
        "Nome_Treinamento": [rotulo] * len(matriculas_presentes),
        "Matrícula": [str(m).strip() for m in matriculas_presentes],
        "Nome": nomes,
    })
    df.to_csv(caminho, sep=";", index=False, encoding="utf-8-sig")
    return caminho


def atualizar_dashboard(caminho_excel: str = ARQUIVO_PADRAO) -> None:
    """
    Atualiza a aba DASHBOARD com % presença e indicadores a partir de PRESENÇA e FUNCIONÁRIOS.
    """
    pres = ler_presenca(caminho_excel)
    func = ler_funcionarios(caminho_excel)
    if pres.empty:
        return
    total_registros = len(pres)
    presentes = (pres["Status"] == "Presente").sum() if "Status" in pres.columns else 0
    pct = (100.0 * presentes / total_registros) if total_registros else 0
    linhas = [
        {"Indicador": "% Presença geral", "Valor": f"{pct:.1f}%", "Detalhe": f"{presentes}/{total_registros} registros"},
        {"Indicador": "Pendentes", "Valor": "", "Detalhe": "Compare PRESENÇA com TREINAMENTOS obrigatórios"},
        {"Indicador": "Concluíram todos", "Valor": "", "Detalhe": "Por setor (use tabela dinâmica)"},
    ]
    dash = pd.DataFrame(linhas)
    todas = _carregar_excel_todas_abas(caminho_excel)
    todas["DASHBOARD"] = dash
    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        for nome, df in todas.items():
            df.to_excel(writer, sheet_name=nome, index=False)
